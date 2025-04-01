"""
This script contains the code whereby the formatting of the output was improved.
"""
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def calculate_probabilities(tutor_row, course_name, student_faculty):
    weights = 0

    if pd.notna(tutor_row["Preferably"]) and course_name == tutor_row["Preferably"]:
        weights += 1.0
    if pd.notna(tutor_row["Then"]) and course_name == tutor_row["Then"]:
        weights += 0.75
    if pd.notna(tutor_row["Then.1"]) and course_name == tutor_row["Then.1"]:
        weights += 0.6
    if pd.notna(tutor_row["Then.2"]) and course_name == tutor_row["Then.2"]:
        weights += 0.5
    if pd.notna(tutor_row["Then.3"]) and course_name == tutor_row["Then.3"]:
        weights += 0.4
    if pd.notna(tutor_row["allocate to Faculty"]) and tutor_row["allocate to Faculty"] in student_faculty:
        weights += 0.5
    if pd.notna(tutor_row["Also"]) and tutor_row["Also"] in student_faculty:
        weights += 0.4
    else:
        weights+=0.01
    
    if pd.notna(tutor_row["But never"]) and course_name in tutor_row["But never"]:
        return 0  
    if pd.notna(tutor_row["But never.1"]) and course_name in tutor_row["But never.1"]:
        return 0  
    if pd.notna(tutor_row["But never.2"]) and course_name in tutor_row["But never.2"]:
        return 0  
    if pd.notna(tutor_row["But never.3"]) and course_name in tutor_row["But never.3"]:
        return 0  

    return weights 

def assign_tutors(tutor_df, student_df):
    tutor_allocation = {tutor: [] for tutor in tutor_df["SPR"]}
    tutor_capacity = {row["SPR"]: row["Allocate (N)"] for _, row in tutor_df.iterrows()}
    unallocated_students = []

    for _, student in student_df.iterrows():
        course_name = student["Course Name"]
        student_faculty = student["Faculty/School"]

        eligible_tutors = tutor_df[~tutor_df["But never"].fillna('').str.contains(course_name, na=False)].copy()
        eligible_tutors["Probability"] = eligible_tutors.apply(
            lambda x: calculate_probabilities(x, course_name, student_faculty), axis=1
        )
        eligible_tutors = eligible_tutors[eligible_tutors["Probability"] > 0]

        if eligible_tutors.empty:
            unallocated_students.append(student["Code"])
            continue

        total_weight = eligible_tutors["Probability"].sum()
       
        eligible_tutors["Normalised Probability"] = eligible_tutors["Probability"] / total_weight

        assigned = False
        tutor_choices = eligible_tutors["SPR"].tolist()
        tutor_probabilities = eligible_tutors["Normalised Probability"].tolist()

        while tutor_choices:
            chosen_tutor = np.random.choice(tutor_choices, p=tutor_probabilities)
            if len(tutor_allocation[chosen_tutor]) < tutor_capacity[chosen_tutor]:
                tutor_allocation[chosen_tutor].append((student["Code"], student["Course Name"]))
                assigned = True
                break
            else:
                index = tutor_choices.index(chosen_tutor)
                tutor_choices.pop(index)
                tutor_probabilities.pop(index)
                if tutor_probabilities:
                    tutor_probabilities = [p / sum(tutor_probabilities) for p in tutor_probabilities]

        if not assigned:
            unallocated_students.append(student["Code"])

    return tutor_allocation, unallocated_students


def rename_duplicates(columns):
    seen = {}
    for i, col in enumerate(columns):
        if col in seen:
            seen[col] += 1
            columns[i] = f"{col}_{seen[col]}"
        else:
            seen[col] = 0
    return columns

"""
This class shwos a slsigth change from original GUI code.
"""
class AssignmentApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Tutor-Student Assignment")

        self.tutor_file = ""
        self.student_file = ""
        self.assignment_result = None
        self.unallocated_students = []

        self.tutor_df = None
        self.student_df = None

        tk.Button(root, text="Upload Tutors CSV", command=self.upload_tutors).pack(pady=10)
        tk.Button(root, text="Upload Students CSV", command=self.upload_students).pack(pady=10)
        tk.Button(root, text="Run Assignment", command=self.run_assignment).pack(pady=10)
        tk.Button(root, text="Download Results", command=self.download_results).pack(pady=10)

    def upload_tutors(self):
        self.tutor_file = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        self.tutor_df = pd.read_csv(self.tutor_file)
        self.tutor_df.columns = rename_duplicates(self.tutor_df.columns)  
        print("Renamed Columns:", self.tutor_df.columns)     
        messagebox.showinfo("File Selected", f"Tutors file selected: {self.tutor_file}")

    def upload_students(self):
        self.student_file = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        self.student_df = pd.read_csv(self.student_file)
        messagebox.showinfo("File Selected", f"Students file selected: {self.student_file}")

    def run_assignment(self):
        if self.tutor_df is None or self.student_df is None:
            messagebox.showwarning("Input Missing", "Please upload both tutor and student CSV files.")
            return

        self.assignment_result, self.unallocated_students = assign_tutors(self.tutor_df, self.student_df)

        if self.unallocated_students:
            messagebox.showinfo("Unallocated Students", f"Some students couldn't be allocated: {self.unallocated_students}")
        else:
            messagebox.showinfo("Assignment Completed", "All students have been successfully allocated.")

    def download_results(self):
        if not self.assignment_result:
            messagebox.showwarning("No Data", "No results to download. Run the assignment first.")
            return

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            wb = Workbook()
            ws_allocated = wb.active
            ws_allocated.title = "Allocated Students"
            ws_allocated.append(["Tutor Name", "Student Code", "Course Name"])

            faculty_colors = {
                'EMS': 'C6EFCE',
                'AHSS': 'FCE4D6',
                'HS': 'D9E1F2'
            }

            for tutor, students in self.assignment_result.items():
                tutor_info = self.tutor_df[self.tutor_df["SPR"] == tutor].iloc[0]
                tutor_name = tutor_info["NAME"]
                faculty = tutor_info["allocate to Faculty"]
                fill = PatternFill(start_color=faculty_colors.get(faculty, 'FFFFFF'), fill_type="solid")

                ws_allocated.append([tutor_name])
                ws_allocated.cell(row=ws_allocated.max_row, column=1).fill = fill

                for student_code, course_name in students:
                    ws_allocated.append(["", student_code, course_name])

            # add a new sheet for unallocated students
            ws_unallocated = wb.create_sheet(title="Unallocated Students")
            ws_unallocated.append(["Student Code", "Course Name"])

            for student_code in self.unallocated_students:
                student_info = self.student_df[self.student_df["Code"] == student_code].iloc[0]
                ws_unallocated.append([student_code, student_info["Course Name"]])

            wb.save(save_path)
            messagebox.showinfo("File Saved", f"Results saved to {save_path}")



if __name__ == "__main__":
    root = tk.Tk()
    app = AssignmentApp(root)
    root.mainloop()

