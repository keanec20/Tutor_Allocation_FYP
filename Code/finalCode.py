"""
This script coantains the final version of the code.
This is what was compiled into a functioning executbale for the Senior Tutor's Office.
True probabilistic approach.

Each function has a brief comment before hand.
"""
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PIL import Image, ImageTk  

"""
This function weights the tutors according to preferences.
Input parameters: data extracted from tutor & student files.
"""
def calculate_probabilities(tutor_row, course_name, student_faculty):
    weights = 0

    if pd.notna(tutor_row["Preferably"]) and course_name == tutor_row["Preferably"]:
        weights += 1.0
    if pd.notna(tutor_row["Then"]) and course_name == tutor_row["Then"]:
        weights += 0.75
    if pd.notna(tutor_row["Then.1"]) and course_name == tutor_row["Then.1"]:
        weights += 0.6
    if pd.notna(tutor_row["Then.2"]) and course_name == tutor_row["Then.2"]:
        weights += 0.5
    if pd.notna(tutor_row["Then.3"]) and course_name == tutor_row["Then.3"]:
        weights += 0.4
    if pd.notna(tutor_row["allocate to Faculty"]) and tutor_row["allocate to Faculty"] in student_faculty:
        weights += 0.5
    if pd.notna(tutor_row["Also"]) and tutor_row["Also"] in student_faculty:
        weights += 0.4
    else:
        weights += 0.01
    
    if pd.notna(tutor_row["But never"]) and course_name in tutor_row["But never"]:
        return 0  
    if pd.notna(tutor_row["But never.1"]) and course_name in tutor_row["But never.1"]:
        return 0  
    if pd.notna(tutor_row["But never.2"]) and course_name in tutor_row["But never.2"]:
        return 0  
    if pd.notna(tutor_row["But never.3"]) and course_name in tutor_row["But never.3"]:
        return 0  

    return weights 

"""
This function is the allocation algorithm.
Calls fucntion to calculate probability weightings. Normalises these to create distribution.
Random selection of tutor based on this. 
Porgress bar updates on each iteration->further comments. 
Input paraemters: student dataframe, tutor dataframe, progress tracker for progress bar.
"""
def assign_tutors(tutor_df, student_df, update_progress=None):
    tutor_allocation = {tutor: [] for tutor in tutor_df["SPR"]}
    tutor_capacity = {row["SPR"]: row["Allocate (N)"] for _, row in tutor_df.iterrows()}
    unallocated_students = []
    
    # track progress
    total_students = len(student_df)
    
    for i, (_, student) in enumerate(student_df.iterrows()):
        # prog uodate
        if update_progress is not None:
            progress_value = (i / total_students) * 100
            update_progress(progress_value)
    
        course_name = student["Course Name"]
        student_faculty = student["Faculty/School"]

        eligible_tutors = tutor_df[~tutor_df["But never"].fillna('').str.contains(course_name, na=False)].copy()
        eligible_tutors["Probability"] = eligible_tutors.apply(
            lambda x: calculate_probabilities(x, course_name, student_faculty), axis=1
        )
        eligible_tutors = eligible_tutors[eligible_tutors["Probability"] > 0]

        if eligible_tutors.empty:
            unallocated_students.append(student["Code"])
            continue

        total_weight = eligible_tutors["Probability"].sum()
       
        eligible_tutors["Normalised Probability"] = eligible_tutors["Probability"] / total_weight

        assigned = False
        tutor_choices = eligible_tutors["SPR"].tolist()
        tutor_probabilities = eligible_tutors["Normalised Probability"].tolist()

        while tutor_choices:
            chosen_tutor = np.random.choice(tutor_choices, p=tutor_probabilities)
            if len(tutor_allocation[chosen_tutor]) < tutor_capacity[chosen_tutor]:
                tutor_allocation[chosen_tutor].append((student["Code"], student["Course Name"]))
                assigned = True
                break
            else:
                index = tutor_choices.index(chosen_tutor)
                tutor_choices.pop(index)
                tutor_probabilities.pop(index)
                if tutor_probabilities:
                    tutor_probabilities = [p / sum(tutor_probabilities) for p in tutor_probabilities]

        if not assigned:
            unallocated_students.append(student["Code"])
    
    # complete progress bar
    if update_progress is not None:
        update_progress(100)

    return tutor_allocation, unallocated_students

"""
This function was written to rename duplicate columns on data extraction from files as there 
were multiple "Then" and "But never" columns.
"""
def rename_duplicates(columns):
    seen = {}
    for i, col in enumerate(columns):
        if col in seen:
            seen[col] += 1
            columns[i] = f"{col}_{seen[col]}"
        else:
            seen[col] = 0
    return columns

"""
Class initiated for the pop-up/interface.
"""
class AssignmentApp:
    """
    This function designs the interface using Tkinter library.
    """
    def __init__(self, root):
        self.root = root
        self.root.title("Tutor-Student Assignment")
        self.root.geometry("600x520")  

        # file tracking
        self.tutor_file = ""
        self.student_file = ""
        self.assignment_result = None
        self.unallocated_students = []
        self.tutor_df = None
        self.student_df = None

        self.main_frame = tk.Frame(root, padx=10, pady=10)
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        self.logo_frame = tk.Frame(self.main_frame, height=80, width=400)
        self.logo_frame.pack(pady=5)
        
        self.logo_label = tk.Label(
            self.logo_frame, 
            text="TUTOR ALLOCATION SYSTEM", 
            font=("Arial", 16, "bold"),
            fg="#4a86e8"
        )
        self.logo_label.pack(pady=10)
        
        
        # file upload
        self.file_frame = tk.LabelFrame(self.main_frame, text="Input Files", padx=10, pady=10)
        self.file_frame.pack(fill=tk.X, pady=5)
        
        # Tutor file 
        tk.Label(self.file_frame, text="Tutors File:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.tutor_path_var = tk.StringVar()
        tk.Entry(self.file_frame, textvariable=self.tutor_path_var, width=40).grid(row=0, column=1, padx=5, pady=5)
        tk.Button(self.file_frame, text="Browse", command=self.upload_tutors).grid(row=0, column=2, padx=5, pady=5)
        
        # Student file 
        tk.Label(self.file_frame, text="Students File:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.student_path_var = tk.StringVar()
        tk.Entry(self.file_frame, textvariable=self.student_path_var, width=40).grid(row=1, column=1, padx=5, pady=5)
        tk.Button(self.file_frame, text="Browse", command=self.upload_students).grid(row=1, column=2, padx=5, pady=5)
        
        # progress info
        self.progress_frame = tk.LabelFrame(self.main_frame, text="Progress", padx=10, pady=10)
        self.progress_frame.pack(fill=tk.X, pady=5)
        
        # progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            self.progress_frame, 
            variable=self.progress_var, 
            mode='determinate',
            length=400
        )
        self.progress_bar.pack(fill=tk.X, pady=5)
        
        # status label
        self.status_var = tk.StringVar(value="Ready to start")
        tk.Label(self.progress_frame, textvariable=self.status_var).pack(anchor=tk.W, pady=5)
        
        # buttons
        self.button_frame = tk.Frame(self.main_frame)
        self.button_frame.pack(fill=tk.X, pady=10)
        
        #run on click
        self.run_btn = tk.Button(
            self.button_frame, 
            text="Allocate Tutors", 
            command=self.run_assignment,
            bg="#4a86e8", 
            fg="white",
            padx=10, 
            pady=5
        )
        self.run_btn.pack(side=tk.LEFT, padx=10)
        
        #downlaod on click
        self.download_btn = tk.Button(
            self.button_frame, 
            text="Download Results", 
            command=self.download_results,
            padx=10, 
            pady=5,
            state=tk.DISABLED
        )
        self.download_btn.pack(side=tk.LEFT, padx=10)
        
    """
    This functgion updates the progress bar during algorithm iterations.
    Input parameters: current progress value.
    """
    def update_progress(self, value):
        self.progress_var.set(value)
        self.root.update_idletasks()

    """
    This function handles potential errors in uploaded data file content.
    Checks for required columns, duplicates, adn missing values.
    Stops user from running algorithm if file doesn't pass checks.
    """
    def validate_data(self):
        #makes sure two files uploaded
        if self.tutor_df is None or self.student_df is None:
            messagebox.showwarning("Input Missing", "Please upload both tutor and student CSV files.")
            return False
        
        # checks for duplicate tutors based on SPR
        duplicate_tutors = self.tutor_df["SPR"].duplicated()
        if duplicate_tutors.any():
            dup_tutor_ids = self.tutor_df.loc[duplicate_tutors, "SPR"].tolist()
            result = messagebox.askyesno("Duplicate Tutors", 
                                        f"Found {len(dup_tutor_ids)} duplicate tutor records. "
                                        f"Only the first occurrence of each will be used. Continue?")
            if not result:
                return False
            # removes duplicates, keeping first appearance of SPR
            self.tutor_df = self.tutor_df.drop_duplicates(subset=["SPR"], keep="first")
            self.status_var.set(f"Removed {len(dup_tutor_ids)} duplicate tutor records")
        
        # checks for duplicate students
        duplicate_students = self.student_df["Code"].duplicated()
        if duplicate_students.any():
            dup_student_ids = self.student_df.loc[duplicate_students, "Code"].tolist()
            result = messagebox.askyesno("Duplicate Students", 
                                        f"Found {len(dup_student_ids)} duplicate student records. "
                                        f"Only the first occurrence of each will be used. Continue?")
            if not result:
                return False
            # removes duplicates, keeping first appearance of student Code
            self.student_df = self.student_df.drop_duplicates(subset=["Code"], keep="first")
            self.status_var.set(f"Removed {len(dup_student_ids)} duplicate student records")
        
        # checks for empty values in required fields-- uses lists created earlier for required columns
        required_student_columns = ["Code", "Course Name", "Faculty/School"]
        required_tutor_columns = ["SPR", "NAME", "Allocate (N)"]
        missing_tutor_data = self.tutor_df[required_tutor_columns].isnull().any(axis=1)
        if missing_tutor_data.any():
            tutor_count = missing_tutor_data.sum()
            result = messagebox.askyesno("Missing Tutor Data", 
                                        f"Found {tutor_count} tutor records with missing required data. "
                                        f"Remove these records and continue?")
            if not result:
                return False
            self.tutor_df = self.tutor_df[~missing_tutor_data]
            self.status_var.set(f"Removed {tutor_count} tutor records with missing data")
        
        missing_student_data = self.student_df[required_student_columns].isnull().any(axis=1)
        if missing_student_data.any():
            student_count = missing_student_data.sum()
            result = messagebox.askyesno("Missing Student Data", 
                                        f"Found {student_count} student records with missing required data. "
                                        f"These students won't be allocated. Continue?")
            if not result:
                return False
            # user asked if they want to remove tutors, whereas studnets just marked as unnallocated
        
        return True

    """
    This function handles the tutor file on upload.
    Carries out inital checks of file format, making sure file has correct columns and isn't empty.
    """
    def upload_tutors(self):
        self.tutor_file = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if self.tutor_file:
            self.tutor_path_var.set(self.tutor_file)
            try:
                self.tutor_df = pd.read_csv(self.tutor_file)
                
                # empty file check
                if self.tutor_df.empty:
                    messagebox.showerror("Error", "The tutor file is empty.")
                    self.tutor_path_var.set("")
                    self.tutor_df = None
                    return
                    
                self.tutor_df.columns = rename_duplicates(self.tutor_df.columns)
                
                # required columns check
                required_columns = ["SPR", "NAME", "Allocate (N)"]
                missing_columns = [col for col in required_columns if col not in self.tutor_df.columns]
                
                if missing_columns:
                    messagebox.showerror("Invalid File Format", 
                                    f"Missing required columns in tutor file: {', '.join(missing_columns)}")
                    self.tutor_path_var.set("")
                    self.tutor_df = None
                    return
                    
                self.status_var.set(f"Tutors file loaded: {os.path.basename(self.tutor_file)}")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load tutor file: {str(e)}")
                self.tutor_path_var.set("")
                self.tutor_df = None

    """
    This function handles the students file on upload.
    Carries out inital checks of file format, making sure file has correct columns and isn't empty.
    """
    def upload_students(self):
        self.student_file = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if self.student_file:
            self.student_path_var.set(self.student_file)
            try:
                self.student_df = pd.read_csv(self.student_file)
                
                # empty file check
                if self.student_df.empty:
                    messagebox.showerror("Error", "The student file is empty.")
                    self.student_path_var.set("")
                    self.student_df = None
                    return
                    
                # required columns check
                required_columns = ["Code", "Course Name", "Faculty/School"]
                missing_columns = [col for col in required_columns if col not in self.student_df.columns]
                
                if missing_columns:
                    messagebox.showerror("Invalid File Format", 
                                    f"Missing required columns in student file: {', '.join(missing_columns)}")
                    self.student_path_var.set("")
                    self.student_df = None
                    return
                    
                self.status_var.set(f"Students file loaded: {os.path.basename(self.student_file)}")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load student file: {str(e)}")
                self.student_path_var.set("")
                self.student_df = None

    """
    This fucntion executes allocation using uploaded files.
    Calls the other fucntions to update progress bar, handle errors, and allows for download of results.
    """
    def run_assignment(self):
        if not self.validate_data():
            return

        self.progress_var.set(0)
        self.status_var.set("Running allocation algorithm...")
        self.root.update()
        
        try:
            self.assignment_result, self.unallocated_students = assign_tutors(
                self.tutor_df, 
                self.student_df,
                self.update_progress
            )
            
            # deals with students with missing data, assigns them to unallocated list
            required_student_columns = ["Code", "Course Name", "Faculty/School"]
            missing_data_mask = self.student_df[required_student_columns].isnull().any(axis=1)
            missing_data_students = self.student_df.loc[missing_data_mask, "Code"].tolist()
            self.unallocated_students.extend([s for s in missing_data_students if s not in self.unallocated_students])
        
            self.status_var.set("Assignment completed")
            
            # download button becomes live
            self.download_btn.config(state=tk.NORMAL)
            
            if self.unallocated_students:
                messagebox.showinfo("Assignment Completed", 
                                f"Assignment completed with {len(self.unallocated_students)} unallocated students.")
            else:
                messagebox.showinfo("Assignment Completed", "All students have been successfully allocated.")
                
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during assignment: {str(e)}")
            self.status_var.set("Allocation failed")

    """
    This function creates an Excel workbook with two seperate sheets for allocated and unallocated students.
    """
    def download_results(self):
            if not self.assignment_result:
                messagebox.showwarning("No Data", "No results to download. Run the assignment first.")
                return

            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                    filetypes=[("Excel files", "*.xlsx")])
            if save_path:
                #update UI status
                self.status_var.set("Generating Excel report...")
                self.root.update()
                
                try:
                    wb = Workbook()
                    ws_allocated = wb.active
                    ws_allocated.title = "Allocated Students"
                    ws_allocated.append(["Tutor Name", "SPR", "Student Code", "Course Name"])

                    faculty_colors = {
                        'EMS': 'C6EFCE',
                        'AHSS': 'FCE4D6',
                        'HS': 'D9E1F2'
                    }

                    for tutor, students in self.assignment_result.items():
                        tutor_info = self.tutor_df[self.tutor_df["SPR"] == tutor].iloc[0]
                        tutor_name = tutor_info["NAME"]
                        faculty = tutor_info["allocate to Faculty"] if pd.notna(tutor_info["allocate to Faculty"]) else "Unknown"
                        fill = PatternFill(start_color=faculty_colors.get(faculty, 'FFFFFF'), fill_type="solid")

                        ws_allocated.append([tutor_name, tutor, "", ""])
                        ws_allocated.cell(row=ws_allocated.max_row, column=1).fill = fill
                        ws_allocated.cell(row=ws_allocated.max_row, column=2).fill = fill

                        for student_code, course_name in students:
                            ws_allocated.append(["", "", student_code, course_name])

                    # sheet for unallocated students
                    ws_unallocated = wb.create_sheet(title="Unallocated Students")
                    ws_unallocated.append(["Student Code", "Course Name"])

                    for student_code in self.unallocated_students:
                        try:
                            student_info = self.student_df[self.student_df["Code"] == student_code].iloc[0]
                            ws_unallocated.append([student_code, student_info["Course Name"]])
                        except (IndexError, KeyError):
                            ws_unallocated.append([student_code, "Unknown"])

                    wb.save(save_path)
                    self.status_var.set(f"Results saved to {os.path.basename(save_path)}")
                    messagebox.showinfo("File Saved", f"Results saved to {save_path}")
                except Exception as e:
                    messagebox.showerror("Error", f"An error occurred while saving the file: {str(e)}")
                    self.status_var.set("File save failed")


"""
RUN THE APP
"""
if __name__ == "__main__":
    root = tk.Tk()
    app = AssignmentApp(root)
    root.mainloop()
